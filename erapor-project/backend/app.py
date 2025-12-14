import io
import os
import re
import zipfile
import tempfile
import subprocess
from typing import List, Optional, Any, Dict

from fastapi import FastAPI, Response, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfMerger

# --- KONFIGURASI ---
TEMPLATE_PATH = "template.xlsx"
SHEET_NAME = "LHBTS"

# Baris Mapping (Sesuaikan dengan template Excel Anda jika berubah)
TEMPLATE_MAPEL_ROWS = {
    "Pendidikan Agama Islam": 16,
    "Pendidikan Pancasila": 17,
    "Bahasa Indonesia": 18,
    "Matematika": 19,
    "Ilmu Pengetahuan Alam": 20,
    "Ilmu Pengetahuan Sosial": 21,
    "Bahasa Inggris": 22,
    "Seni Rupa": 23,
    "Pend. Jasmani, Olahraga dan Kesehatan": 24,
    "Informatika": 25,
    "Koding dan Kecerdasan Artifisial": 26,
    "Bahasa Bali": 28,
    "Bahasa Arab": 30,
}

# Mapping Frontend (Uppercase) -> Template
FRONTEND_MAPEL_TO_TEMPLATE = {
    "AGAMA ISLAM": "Pendidikan Agama Islam",
    "PKN": "Pendidikan Pancasila",
    "BAHASA INDONESIA": "Bahasa Indonesia",
    "MATEMATIKA": "Matematika",
    "IPA": "Ilmu Pengetahuan Alam",
    "IPS": "Ilmu Pengetahuan Sosial",
    "BAHASA INGGRIS": "Bahasa Inggris",
    "SENI BUDAYA": "Seni Rupa",
    "OLAHRAGA": "Pend. Jasmani, Olahraga dan Kesehatan",
    "INFORMATIKA": "Informatika",
    "KKA": "Koding dan Kecerdasan Artifisial",
    "BAHASA BALI": "Bahasa Bali",
    "BAHASA ARAB": "Bahasa Arab",
}

FORMATIF_COLS = [chr(c) for c in range(ord("D"), ord("W") + 1)] # D-W
SUMATIF_COLS = ["Y", "Z", "AA", "AB", "AC"]

# --- HELPER FUNCTIONS ---
def safe_sheet_name(s: str) -> str:
    s = re.sub(r"[\[\]\*\/\\\?\:]", " ", s)
    return re.sub(r"\s+", " ", s).strip()[:31] or "Sheet"

def sanitize_filename(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_\-\. ]+", "", name).strip().replace(" ", "_")

def normalize_mapel_key(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).upper()

# --- PYDANTIC MODELS ---
class MapelItem(BaseModel):
    mapel_key: str
    na: Optional[float] = None
    deskripsi: Optional[str] = None
    # Opsional: formatif/sumatif list jika ingin detail
    formatif: Optional[List[float]] = None
    sumatif: Optional[List[float]] = None

class SiswaPayload(BaseModel):
    no_urut: int = 1
    nama_peserta: str
    nis_nisn: str = "-"
    nama_sekolah: str = "SMP ALBANNA"
    alamat_sekolah: str = "Jl. Tukad Yeh Ho III No. 16"
    kelas: str
    fase: str = "D"
    semester: str = "1 (Ganjil)"
    tahun_ajaran: str = "2024/2025"
    mapel: List[MapelItem] = Field(default_factory=list)

class BatchPayload(BaseModel):
    siswa: List[SiswaPayload]

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def root():
    return {"message": "E-Rapor API is Running", "docs": "/docs"}

# --- EXCEL LOGIC ---
def apply_page_setup(ws: Worksheet):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9 # A4

def fill_student_data(ws: Worksheet, s: SiswaPayload):
    # Identitas
    ws["A2"].value = s.no_urut
    ws["D6"].value = f": {s.nama_peserta}"
    ws["D7"].value = f": {s.nis_nisn}"
    ws["D8"].value = f": {s.nama_sekolah}"
    ws["D9"].value = f": {s.alamat_sekolah}"
    ws["AF6"].value = f": {s.kelas}"
    ws["AF7"].value = f": {s.fase}"
    ws["AF8"].value = f": {s.semester}"
    ws["AF9"].value = f": {s.tahun_ajaran}"

    # Bersihkan Nilai Lama
    for row in TEMPLATE_MAPEL_ROWS.values():
        for col in FORMATIF_COLS + SUMATIF_COLS + ["X", "AD", "AE", "AF"]:
            ws[f"{col}{row}"].value = None

    # Isi Nilai Baru
    total_na = 0; count_na = 0
    for item in s.mapel:
        key = normalize_mapel_key(item.mapel_key)
        tpl_name = FRONTEND_MAPEL_TO_TEMPLATE.get(key)
        if not tpl_name or tpl_name not in TEMPLATE_MAPEL_ROWS: continue
        
        row = TEMPLATE_MAPEL_ROWS[tpl_name]
        
        # Nilai Akhir & Deskripsi
        if item.na is not None:
            ws[f"AE{row}"].value = item.na
            total_na += item.na
            count_na += 1
        if item.deskripsi:
            ws[f"AF{row}"].value = item.deskripsi

    # Rata-rata Total
    if count_na > 0:
        ws["AE31"].value = total_na
        ws["AE32"].value = total_na / count_na

def create_workbook_for_student(s: SiswaPayload) -> Any:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]
    apply_page_setup(ws)
    fill_student_data(ws, s)
    return wb

# --- ENDPOINTS ---

@app.post("/api/rapor/xlsx")
def generate_xlsx(payload: BatchPayload, combine: int = Query(0)):
    if not payload.siswa: return Response("Data kosong", status_code=400)

    # Mode 1: Combine (Satu File Excel Banyak Sheet)
    if combine == 1:
        wb_master = load_workbook(TEMPLATE_PATH)
        ws_master = wb_master[SHEET_NAME]
        
        # Isi sheet pertama
        fill_student_data(ws_master, payload.siswa[0])
        ws_master.title = safe_sheet_name(payload.siswa[0].nama_peserta)

        # Tambah sheet siswa lain
        for s in payload.siswa[1:]:
            ws_copy = wb_master.copy_worksheet(ws_master) # Copy layout
            # Reset values karena copy_worksheet mengcopy value juga
            fill_student_data(ws_copy, s) 
            ws_copy.title = safe_sheet_name(s.nama_peserta)
        
        out = io.BytesIO()
        wb_master.save(out)
        fname = f"Rapor_Gabungan_{payload.siswa[0].kelas}.xlsx"
        return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    # Mode 2: Batch ZIP (Banyak File Excel) atau Single File
    if len(payload.siswa) == 1:
        wb = create_workbook_for_student(payload.siswa[0])
        out = io.BytesIO()
        wb.save(out)
        fname = sanitize_filename(f"Rapor_{payload.siswa[0].nama_peserta}.xlsx")
        return Response(out.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    
    # ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in payload.siswa:
            wb = create_workbook_for_student(s)
            buf = io.BytesIO()
            wb.save(buf)
            zf.writestr(sanitize_filename(f"Rapor_{s.nama_peserta}.xlsx"), buf.getvalue())
    
    return Response(zip_buffer.getvalue(), media_type="application/zip", headers={"Content-Disposition": 'attachment; filename="Rapor_Batch_XLSX.zip"'})


@app.post("/api/rapor/pdf")
def generate_pdf(payload: BatchPayload, combine: int = Query(0)):
    if not payload.siswa: return Response("Data kosong", status_code=400)

    def convert_to_pdf(wb_obj) -> bytes:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = os.path.join(tmp, "temp.xlsx")
            wb_obj.save(xlsx_path)
            # LibreOffice Command
            subprocess.run(["soffice", "--headless", "--convert-to", "pdf", "--outdir", tmp, xlsx_path], check=True)
            with open(os.path.join(tmp, "temp.pdf"), "rb") as f:
                return f.read()

    # Mode 1: Single PDF
    if len(payload.siswa) == 1 and combine == 0:
        wb = create_workbook_for_student(payload.siswa[0])
        pdf_bytes = convert_to_pdf(wb)
        fname = sanitize_filename(f"Rapor_{payload.siswa[0].nama_peserta}.pdf")
        return Response(pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    # Mode 2: Combine PDF (Merge)
    if combine == 1:
        merger = PdfMerger()
        for s in payload.siswa:
            wb = create_workbook_for_student(s)
            pdf_bytes = convert_to_pdf(wb)
            merger.append(io.BytesIO(pdf_bytes))
        
        out = io.BytesIO()
        merger.write(out)
        fname = f"Rapor_Gabungan_{payload.siswa[0].kelas}.pdf"
        return Response(out.getvalue(), media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    # Mode 3: ZIP PDFs
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for s in payload.siswa:
            wb = create_workbook_for_student(s)
            pdf_bytes = convert_to_pdf(wb)
            zf.writestr(sanitize_filename(f"Rapor_{s.nama_peserta}.pdf"), pdf_bytes)
    
    return Response(zip_buffer.getvalue(), media_type="application/zip", headers={"Content-Disposition": 'attachment; filename="Rapor_Batch_PDF.zip"'})
```

---

### TAHAP 2: Upload ke GitHub

1.  Buka [GitHub.com](https://github.com) dan buat **Repository Baru** (misal: `erapor-backend`). Jangan centang "Add README" agar repository kosong.
2.  Buka terminal/command prompt di komputer Anda.
3.  Arahkan ke folder `erapor-project` yang baru Anda buat.
4.  Jalankan perintah berikut satu per satu:

```bash
git init
git add .
git commit -m "Upload pertama backend erapor"
git branch -M main
# Ganti URL di bawah dengan URL repo GitHub Anda sendiri
git remote add origin https://github.com/USERNAME_ANDA/erapor-backend.git
git push -u origin main